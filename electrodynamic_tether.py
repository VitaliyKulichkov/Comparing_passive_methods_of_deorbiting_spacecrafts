from abc import ABC
import mpmath
import openpyxl
import matplotlib.pyplot as plt
import pandas as pd


class Tether(ABC):
    """
    Class of Tether with constants
    """

    def __init__(self):
        """
        Construct
        """
        # height of a density layers of atmosphere
        self.height_of_density_layers_of_atmosphere = 120000
        # value of magnetic induction at the geomagnetic equator
        self.bm = 4 * pow(10, -5)
        # value of wire rope resistance
        self.wire_rope_resistance = 125 * pow(10, -8)
        # constant for finding bz
        self.mu0 = 1.25663706212 * pow(10, -6)
        # value of magnetic dipole moment
        self.pm = 8.3 * pow(10, 22)
        self.r_square = None
        self.y = None
        self.x = None
        self.height_of_orbit = None
        self.weight_system = None
        self.weight_of_tether = None

    @staticmethod
    def print_init():
        print("Initialized")

    @staticmethod
    def print_start():
        print("Start calculation")


class TetherSon(Tether):
    """
    Class representing a Tether.
    """

    def __init__(self):
        """
        Construct
        """
        super(TetherSon, self).__init__()
        self.delta_to_months = None
        self.deltat = None
        self.kepler_a = None
        self.coskvim = None
        self.cos = None
        self.btz = None
        self.bz = None

    def weight_of_system(self, length, diametr, mka) -> float:
        """
        This function calculating weight of system of spacecraft and tether system
        :param length: length of tether system
        :param diametr: diametr of tether system
        :param mka: weight of spacecraft
        :return: weight of system of spacecraft and tether system
        """
        self.weight_of_tether = 7660 * mpmath.pi * diametr * diametr * length / 4
        self.weight_system = mka + self.weight_of_tether
        return self.weight_system

    def coordinates_of_spacecraft(self, perigee, radius_earth=6371000) -> tuple:
        """
        This function calculating coordinates of spacecraft
        :param perigee: Height of the orbit
        :param radius_earth: Radius of Earth
        :return: coordinate "X" and "Y" of the spacecraft
        """
        self.height_of_orbit = radius_earth + perigee
        self.x = 0.5 * self.height_of_orbit
        self.y = mpmath.sqrt(self.height_of_orbit * self.height_of_orbit - self.x * self.x)
        coordinates = (self.x, self.y)
        return coordinates

    def magnetic_induction(self, x, y, kepler_i) -> float:
        """
        This function calculate tilt of the earth's magnetic field axis and return
        value of function (here cos()) depending on
        tilt of the earth's magnetic field axis.
        :param x: coordinate "X" of the spacecraft
        :param y: coordinate "Y" of the spacecraft
        :param kepler_i: orbital inclination
        :return: value of function (here cos()) depending on
        tilt of the earth's magnetic field axis.
        """
        self.r_square = x * x + y * y
        self.bz = -self.mu0 * (1 / (4 * mpmath.pi)) * (self.pm /
                                                       (self.r_square * mpmath.sqrt(self.r_square)))
        self.btz = self.bz
        self.cos = (self.btz * self.bz) / (self.btz * self.bz)
        self.coskvim = (
                6 + 2 * mpmath.cos(2 * kepler_i) + 3 * mpmath.cos(2 * kepler_i) + 2 * self.cos
                + 3 * mpmath.cos(2 * kepler_i))
        return self.coskvim

    def deorbiting_time(self, weight_system, length, coskvim, perigee, apogee) -> float:
        """
        This function returns time of deorbiting the spacecraft
        by using tether system
        :param apogee: apogee of an orbit
        :param weight_system: weight of system of spacecraft and tether system
        :param length: length of tether system
        :param coskvim: value of function (here cos())
        depending on tilt of the earth's magnetic field axis.
        :param perigee: Height which from we deorbiting our spacecraft
        :return: Time of deorbiting the spacecraft
        """
        self.kepler_a = (apogee + perigee) / 2
        self.deltat = ((weight_system * self.wire_rope_resistance) / (2 * length * pow(self.kepler_a, 6) *
                                                                      1 * coskvim * self.bm * self.bm)) * (
                              pow(perigee, 7) - pow(self.height_of_density_layers_of_atmosphere, 7))
        self.delta_to_months = self.deltat * 3.8052 * pow(10, -7)
        print(
            f"Время увода системы массой {weight_system} с высоты орбиты {perigee}"
            f" тросовой системой длинной {length} составляет {self.delta_to_months} мес.")
        return self.delta_to_months

    @staticmethod
    def write_data_into_excel(lst: list) -> None:
        """
        This function adding data with results to Excel sheets
        :param lst: 2 dimension list with results
        :return: Excel file with inserted data
        """
        sheets = ['L-5 000 M-900', 'L-10 000 M-900', 'L-15 000 M-900', 'L-20 000 M-900', 'L-25 000 M-900',
                  'L-5 000 M-1 800', 'L-10 000 M-1 800', 'L-15 000 M-1 800', 'L-20 000 M-1 800', 'L-25 000 M-1 800']
        title = ["Высота орбиты, м", "Масса КА, кг", "Масса системы, кг", "Длина троса, м", "Время увода, мес"]
        book = openpyxl.Workbook()
        for sheet in sheets:
            book.create_sheet(sheet)
            book[sheet].append(title)
        for i in lst:
            if i[1] == 900 and i[3] == 5000:
                book['L-5 000 M-900'].append(i)
            elif i[1] == 900 and i[3] == 10000:
                book['L-10 000 M-900'].append(i)
            elif i[1] == 900 and i[3] == 15000:
                book['L-15 000 M-900'].append(i)
            elif i[1] == 900 and i[3] == 20000:
                book['L-20 000 M-900'].append(i)
            elif i[1] == 900 and i[3] == 25000:
                book['L-25 000 M-900'].append(i)
            elif i[1] == 1800 and i[3] == 5000:
                book['L-5 000 M-1 800'].append(i)
            elif i[1] == 1800 and i[3] == 10000:
                book['L-10 000 M-1 800'].append(i)
            elif i[1] == 1800 and i[3] == 15000:
                book['L-15 000 M-1 800'].append(i)
            elif i[1] == 1800 and i[3] == 20000:
                book['L-20 000 M-1 800'].append(i)
            elif i[1] == 1800 and i[3] == 25000:
                book['L-25 000 M-1 800'].append(i)
        book.save('tether2.xlsx')

    @staticmethod
    def creating_a_diagram():
        """
        This function create plots and inserting it into Excel File with results
        :return: Excel file with diagrams
        """
        sheets = ['L-5 000 M-900', 'L-10 000 M-900', 'L-15 000 M-900', 'L-20 000 M-900', 'L-25 000 M-900',
                  'L-5 000 M-1 800', 'L-10 000 M-1 800', 'L-15 000 M-1 800', 'L-20 000 M-1 800', 'L-25 000 M-1 800']
        wb = openpyxl.load_workbook('tether2.xlsx')
        for sheet in sheets:
            ws = wb[sheet]
            var = pd.read_excel("tether2.xlsx", sheet_name=sheet)
            x = list(var['Время увода, мес'])
            y = list(var['Высота орбиты, м'])
            plt.figure(figsize=(7, 5))
            plt.scatter(x, y, marker=".", s=100, edgecolors="black", c="yellow")
            plt.title(sheet)
            plt.xlabel('Deorbiting time, months', fontweight='bold', color='black',
                       fontsize='12', horizontalalignment='center')
            plt.ylabel('Perigee, m.', fontweight='bold', color='black', fontsize='12', horizontalalignment='center')
            plt.plot(x, y, '-o')
            # plt.show()
            plt.savefig(f'{sheet}.png')
            img = openpyxl.drawing.image.Image(f'{sheet}.png')
            img.anchor = 'G1'
            ws.add_image(img)
            wb.save('tether2.xlsx')

    def calculate_all_results(self) -> None:
        """
        This function calculating all parameters, inserting it in Excel file, creating diagrams and inserting it in
        Excel file.
        :return: Excel file with calculated parameters and diagrams.
        """
        ######################################################
        # Initial data, which u can change
        ######################################################
        list_of_weights = [900, 1800]
        list_of_perigee = [700000, 800000, 900000]
        list_of_apogee = [800000, 900000, 1000000]
        length_of_tether = [5000, 10000, 15000, 20000, 25000]
        zipped_heights = list(zip(list_of_apogee, list_of_perigee))
        results = [
            ["Высота орбиты, м", "Масса КА, кг", "Масса системы, кг", "Длина троса, м", "Время увода, мес"],
        ]
        for length in length_of_tether:
            for mass in list_of_weights:
                for tuple_of_heights in zipped_heights:
                    weight = self.weight_of_system(length, 0.007, mass)
                    coordinates = self.coordinates_of_spacecraft(tuple_of_heights[0])
                    induction = self.magnetic_induction(coordinates[0], coordinates[1], 60)
                    time = float(
                        self.deorbiting_time(weight, length, induction, tuple_of_heights[0], tuple_of_heights[1]))
                    results.append([tuple_of_heights[0], int(mass), int(weight), length, time])
        self.write_data_into_excel(results)
        self.creating_a_diagram()
        print('Result calculated!')


def main() -> None:
    """
    Start calculating
    :return: Excel file
    """
    tether = TetherSon()
    tether.calculate_all_results()


if __name__ == '__main__':
    main()
